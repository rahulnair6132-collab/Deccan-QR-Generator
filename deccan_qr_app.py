#!/usr/bin/env python3
"""
Deccan Enterprises — Dakshata Suite (Web App)
=============================================
Two modules in one Streamlit app, selected from the sidebar:

  1. QR Generator (default)
     Same engine as before: Job Card + MR + Qty -> QR PNGs (Insulator ID
     printed below), ZIP + Excel/CSV manifest.

  2. Production Entry Visualizer
     Upload the daily ScanLog export from SharePoint (.xlsx / .csv;
     .iqy is detected and explained), pick a shift window
     (e.g. 6th 08:00 -> 7th 08:00), and get:
       * ERP-ready Excel report: one sheet per stage with a Job-Card
         summary (Total / OK / Rework / Rejection) + insulator-level
         detail. Rejection & Rework roll into the stage they occurred
         at (StageAsked). Separate "Reused Rejected" sheet.
       * In-app preview of every sheet.
       * Optional analytics dashboard (interactive charts, filters,
         PNG export via the chart camera icon).

     Data-repair built in (both toggleable):
       * Column-shift repair — some SharePoint exports shift the DT
         columns one place left starting at Rejection_MC.
       * Month/day swap fix — timestamps like 2026-10-06 that should
         be 2026-06-10.

RUN:      streamlit run deccan_qr_app.py
INSTALL:  py -3 -m pip install streamlit qrcode pillow openpyxl pandas   (charts use Altair, bundled with Streamlit)
Repo files expected next to this script: DejaVuSans-Bold.ttf, deccan_logo.png
"""

import io
import os
import csv
import zipfile
from datetime import datetime, date, time, timedelta

import streamlit as st
import pandas as pd
import qrcode
from qrcode.constants import ERROR_CORRECT_Q
from PIL import Image, ImageDraw, ImageFont

# ======================================================================
# THEME / BRANDING
# ======================================================================
DECCAN_BLUE = "#1B5EA6"      # primary brand blue (from the logo)
DECCAN_BLUE_DARK = "#134A85"
DECCAN_BLUE_LIGHT = "#E8F0F9"
DECCAN_GREY = "#6B7280"

HERE = os.path.dirname(os.path.abspath(__file__))
LOGO_PATH = os.path.join(HERE, "deccan_logo.png")

STAGES = ["CBM", "Priming", "Moulding", "CAM", "SML", "Cframe", "Packing", "Invoicing"]
STAGE_LABELS = {"CBM": "CBM (Crimping)", "Priming": "Priming", "Moulding": "Moulding",
                "CAM": "CAM", "SML": "SML", "Cframe": "C-Frame", "Packing": "Packing",
                "Invoicing": "Invoicing"}
# machine column names as they appear in the export (note CFrame_MC casing)
MC_COL = {"CBM": "CBM_MC", "Moulding": "Moulding_MC", "CAM": "CAM_MC",
          "SML": "SML_MC", "Cframe": "CFrame_MC"}
DT_COL = {s: f"{s}_DT" for s in STAGES}
# order used for the column-shift repair (values sit one column LEFT of home)
SHIFT_CHAIN = ["Rejection_MC", "CBM_DT", "Priming_DT", "Moulding_DT", "CAM_DT",
               "SML_DT", "Rework_DT", "Cframe_DT", "Packing_DT", "Rejection_DT",
               "Invoicing_DT"]


def inject_css():
    st.markdown(f"""
    <style>
      .stApp {{ background: #FFFFFF; }}
      section[data-testid="stSidebar"] {{
          background: {DECCAN_BLUE_LIGHT};
          border-right: 1px solid #d5e2f0;
      }}
      h1, h2, h3 {{ color: {DECCAN_BLUE_DARK}; }}
      .stButton > button[kind="primary"], .stDownloadButton > button {{
          background: {DECCAN_BLUE}; color: #fff; border: none; border-radius: 6px;
      }}
      .stButton > button[kind="primary"]:hover, .stDownloadButton > button:hover {{
          background: {DECCAN_BLUE_DARK}; color: #fff;
      }}
      div[data-testid="stMetric"] {{
          background: #FFFFFF; border: 1px solid #d5e2f0;
          border-left: 4px solid {DECCAN_BLUE};
          border-radius: 8px; padding: 10px 14px;
          box-shadow: 0 1px 3px rgba(19,74,133,.08);
      }}
      div[data-testid="stMetric"] label {{ color: {DECCAN_GREY}; }}
      div[data-testid="stExpander"] {{
          border: 1px solid #d5e2f0; border-radius: 8px;
      }}
      button[data-baseweb="tab"] {{
          font-weight: 600;
      }}
      thead tr th {{ background: {DECCAN_BLUE_LIGHT} !important; }}
    </style>""", unsafe_allow_html=True)


def sidebar_header():
    with st.sidebar:
        st.markdown(f"<div style='color:{DECCAN_BLUE};font-weight:700;font-size:20px'>"
                    "DAKSHATA SUITE</div>"
                    f"<div style='color:{DECCAN_GREY};font-size:12px'>Deccan Enterprises "
                    "· Since 1966</div>", unsafe_allow_html=True)
        st.markdown("---")


def page_header(title, subtitle):
    """Centered Deccan logo above a clean title block."""
    if os.path.exists(LOGO_PATH):
        l, m, r = st.columns([2.2, 1, 2.2])
        with m:
            st.image(LOGO_PATH, width="stretch")
    else:
        st.markdown(f"<div style='text-align:center'><span style='color:{DECCAN_BLUE};"
                    "font-weight:800;font-size:32px;letter-spacing:1px'>DECCAN</span><br>"
                    f"<span style='color:{DECCAN_BLUE};font-size:11px;letter-spacing:4px'>"
                    "— SINCE 1966 —</span></div>", unsafe_allow_html=True)
    st.markdown(f"""
        <div style='text-align:center;margin:6px 0 4px 0'>
          <div style='font-size:30px;font-weight:700;color:{DECCAN_BLUE_DARK};
                      line-height:1.2'>{title}</div>
          <div style='color:{DECCAN_GREY};font-size:14px;margin-top:4px'>{subtitle}</div>
        </div>
        <hr style='border:none;border-top:2px solid {DECCAN_BLUE_LIGHT};
                   margin:12px 0 18px 0'>""", unsafe_allow_html=True)


# ======================================================================
# QR GENERATOR (engine unchanged, UI restyled)
# ======================================================================
QR_TARGET_PX = 640
CAPTION_FONT_SIZE = 93
FONT_FALLBACK = False


def _load_font(size):
    """Load a bold TrueType font at the requested size.

    Looks FIRST for the DejaVuSans-Bold.ttf bundled next to this script
    (so it works identically on Streamlit Cloud, where system fonts may
    be missing). Falls back to common system fonts, then — only as a last
    resort — Pillow's tiny bitmap default. If we ever hit that last
    resort, FONT_FALLBACK is set True so the app can warn the user.
    """
    here = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(here, "DejaVuSans-Bold.ttf"),   # bundled with the app
        "DejaVuSans-Bold.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "arialbd.ttf",
        r"C:\Windows\Fonts\arialbd.ttf",
        r"C:\Windows\Fonts\arial.ttf",
        r"C:\Windows\Fonts\segoeuib.ttf",
    ]
    for path in candidates:
        try:
            return ImageFont.truetype(path, size)
        except Exception:
            continue
    global FONT_FALLBACK
    FONT_FALLBACK = True
    return ImageFont.load_default()


def make_qr_image(data, box_size=20, border=2):
    qr = qrcode.QRCode(version=None, error_correction=ERROR_CORRECT_Q,
                       box_size=box_size, border=border)
    qr.add_data(data)
    qr.make(fit=True)
    return qr.make_image(fill_color="black", back_color="white").convert("RGB")


def label_with_caption(code_img, caption, sku="", target_px=QR_TARGET_PX, font_size=CAPTION_FONT_SIZE):
    """QR with the Insulator ID printed large below it, and (optionally)
    the SKU on a second centered line beneath the ID. The image widens
    if the text is wider than the QR, so nothing is clipped."""
    code_img = code_img.resize((target_px, target_px), Image.NEAREST)
    font = _load_font(font_size)

    measure = ImageDraw.Draw(Image.new("RGB", (10, 10), "white"))

    def dims(text):
        b = measure.textbbox((0, 0), text, font=font)
        return (b[2] - b[0], b[3] - b[1], b[1])   # width, height, top offset

    lines = [caption]
    if sku:
        lines.append(sku)

    line_dims = [dims(t) for t in lines]
    max_w = max(w for w, _, _ in line_dims)
    line_gap = int(font_size * 0.25)
    pad_above = 16
    pad_below = 22

    total_text_h = sum(h for _, h, _ in line_dims) + line_gap * (len(lines) - 1)
    cap_h = total_text_h + pad_above + pad_below
    canvas_w = max(target_px, max_w + 8)

    canvas = Image.new("RGB", (canvas_w, target_px + cap_h), "white")
    qr_x = (canvas_w - target_px) // 2
    canvas.paste(code_img, (qr_x, 0))
    draw = ImageDraw.Draw(canvas)

    y = target_px + pad_above
    for text, (w, h, top_off) in zip(lines, line_dims):
        draw.text(((canvas_w - w) / 2, y - top_off), text, fill="black", font=font)
        y += h + line_gap
    return canvas


def build_outputs(jobcard, sku, mr_jobs, start=1):
    """Generate all QR images + manifests in memory.

    One SKU applies to the whole job card (not per MR).

    Returns: (rows, images, xlsx_bytes, csv_bytes, zip_bytes)
      rows   : list of dicts (InsulatorID, JobCard, MRNo, SKU, QRPayload)
      images : list of (mr, insulator_id, PIL.Image) for preview
    """
    jobcard = jobcard.strip().upper()
    sku = (sku or "").strip()
    rows = []
    images = []
    n = start

    for mr, qty in mr_jobs:
        mr = mr.strip().upper()
        for _ in range(qty):
            serial = f"{jobcard}-{n:05d}"
            payload = f"{jobcard}|{mr}|{serial}"           # ID + MR in the QR
            img = label_with_caption(make_qr_image(payload), serial, sku=sku)
            images.append((mr, serial, img))
            rows.append({
                "InsulatorID": serial, "JobCard": jobcard, "MRNo": mr,
                "SKU": sku,
                "QRPayload": payload,
                "GeneratedOn": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            })
            n += 1

    # --- Excel manifest (JobCard / MRNo / SKU / InsulatorID) ---
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = "Manifest"
    ws.append(["JobCard", "MRNo", "SKU", "InsulatorID"])
    for r in rows:
        ws.append([r["JobCard"], r["MRNo"], r["SKU"], r["InsulatorID"]])
    for col_letter, width in [("A", 14), ("B", 18), ("C", 16), ("D", 22)]:
        ws.column_dimensions[col_letter].width = width
    xlsx_buf = io.BytesIO()
    wb.save(xlsx_buf)
    xlsx_bytes = xlsx_buf.getvalue()

    # --- CSV manifest (full columns, like the script) ---
    csv_buf = io.StringIO()
    w = csv.DictWriter(csv_buf, fieldnames=list(rows[0].keys()))
    w.writeheader(); w.writerows(rows)
    csv_bytes = csv_buf.getvalue().encode("utf-8")

    # --- ZIP with the JobCard/MR/image.png folder layout ---
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for mr, serial, img in images:
            png_buf = io.BytesIO()
            img.save(png_buf, format="PNG")
            zf.writestr(f"{jobcard}/{mr}/{serial}.png", png_buf.getvalue())
        zf.writestr(f"{jobcard}/{jobcard}_manifest.xlsx", xlsx_bytes)
        zf.writestr(f"{jobcard}/{jobcard}_manifest.csv", csv_bytes)
    zip_bytes = zip_buf.getvalue()

    return rows, images, xlsx_bytes, csv_bytes, zip_bytes


def page_qr_generator():
    page_header("Insulator QR Code Generator",
                "Enter a Job Card and MR numbers, generate shop-floor QR codes "
                "with the Excel manifest.")
    _ = label_with_caption(make_qr_image("TEST"), "TEST")
    if FONT_FALLBACK:
        st.warning("DejaVuSans-Bold.ttf not found next to the script — the "
                   "Insulator ID text will render small. Upload the font to the repo.")

    jc_col, sku_col = st.columns(2)
    jobcard = jc_col.text_input("Job Card No.", placeholder="e.g. JA266-009").strip()
    sku = sku_col.text_input("SKU No.", placeholder="e.g. 1098 A3",
                             help="One SKU applies to the whole job card.").strip()

    st.markdown("##### MR numbers & quantities")
    if "mr_rows" not in st.session_state:
        st.session_state.mr_rows = [{"mr": "", "qty": 10}]

    def add_row():
        st.session_state.mr_rows.append({"mr": "", "qty": 10})

    def remove_row(i):
        if len(st.session_state.mr_rows) > 1:
            st.session_state.mr_rows.pop(i)

    for i, row in enumerate(st.session_state.mr_rows):
        c1, c2, c3 = st.columns([5, 2, 1])
        st.session_state.mr_rows[i]["mr"] = c1.text_input(
            "MR No.", value=row["mr"], key=f"mr_{i}", placeholder="e.g. R126605-005",
            label_visibility="collapsed" if i else "visible")
        st.session_state.mr_rows[i]["qty"] = c2.number_input(
            "Qty", min_value=1, max_value=100000, value=int(row["qty"]), step=1,
            key=f"qty_{i}", label_visibility="collapsed" if i else "visible")
        if i == 0:
            c3.markdown("&nbsp;")
        c3.button("🗑", key=f"del_{i}", on_click=remove_row, args=(i,),
                  disabled=(len(st.session_state.mr_rows) == 1), help="Remove this row")

    st.button("➕ Add another MR", on_click=add_row)

    with st.expander("Advanced options"):
        start_serial = st.number_input(
            "Starting serial number", min_value=1, value=1, step=1,
            help="The first insulator serial. Leave at 1 unless continuing a previous batch.")

    st.divider()
    if st.button("Generate QR codes", type="primary", width="stretch"):
        if not jobcard:
            st.error("Please enter a Job Card No."); st.stop()
        if not sku:
            st.error("Please enter the SKU No. for this job card."); st.stop()
        mr_jobs = [( (r["mr"] or "").strip(), int(r["qty"]) )
                   for r in st.session_state.mr_rows
                   if (r["mr"] or "").strip() and int(r["qty"]) > 0]
        if not mr_jobs:
            st.error("Please enter at least one MR No. with a quantity."); st.stop()
        total = sum(q for _, q in mr_jobs)
        with st.spinner(f"Generating {total} QR code(s)…"):
            rows, images, xlsx_bytes, csv_bytes, zip_bytes = build_outputs(
                jobcard, sku, mr_jobs, start=int(start_serial))
        st.success(f"Generated {len(rows)} QR code(s) across {len(mr_jobs)} MR number(s).")
        st.session_state.result = {"jobcard": jobcard.upper(), "rows": rows,
                                   "images": images, "xlsx": xlsx_bytes,
                                   "csv": csv_bytes, "zip": zip_bytes}

    res = st.session_state.get("result")
    if res:
        jobcard = res["jobcard"]
        st.markdown("### Download")
        d1, d2 = st.columns(2)
        d1.download_button("⬇ All QR codes + manifest (ZIP)", data=res["zip"],
                           file_name=f"{jobcard}_QR_codes.zip", mime="application/zip",
                           width="stretch")
        d2.download_button("⬇ Excel manifest only (.xlsx)", data=res["xlsx"],
                           file_name=f"{jobcard}_manifest.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           width="stretch")
        st.markdown("### Manifest preview")
        st.dataframe([{"JobCard": r["JobCard"], "MRNo": r["MRNo"], "SKU": r["SKU"],
                       "InsulatorID": r["InsulatorID"]} for r in res["rows"]],
                     width="stretch", hide_index=True)
        st.markdown("### QR preview")
        st.caption("Showing the first few QR codes. Download the ZIP for all of them.")
        preview = res["images"][:6]
        cols = st.columns(3)
        for idx, (mr, serial, img) in enumerate(preview):
            buf = io.BytesIO(); img.save(buf, format="PNG")
            cols[idx % 3].image(buf.getvalue(), caption=f"{mr}", width="stretch")
        if len(res["images"]) > len(preview):
            st.caption(f"… and {len(res['images']) - len(preview)} more in the ZIP.")


# ======================================================================
# PRODUCTION ENTRY VISUALIZER — data engine
# ======================================================================
def _looks_like_datetime(v):
    if pd.isna(v):
        return False
    if isinstance(v, (pd.Timestamp, datetime)):
        return True
    s = str(v)
    return ("T" in s and s[:4].isdigit()) or (len(s) >= 10 and s[:4].isdigit() and s[4] == "-")


def repair_shifted_rows(df):
    """Some SharePoint exports shift the timestamp block one column LEFT
    starting at Rejection_MC (Rejection_MC holds CBM_DT, CBM_DT holds
    Priming_DT, … Rejection_DT holds Invoicing_DT). Detect rows where
    Rejection_MC parses as a datetime and shift the chain back RIGHT."""
    cols = [c for c in SHIFT_CHAIN if c in df.columns]
    df["_WasShifted"] = False
    if "Rejection_MC" not in cols:
        return df, 0
    fixed = 0
    for idx in df.index:
        if _looks_like_datetime(df.at[idx, "Rejection_MC"]):
            vals = [df.at[idx, c] for c in cols]
            new_vals = [pd.NA] + vals[:-1]          # shift right by one
            for c, v in zip(cols, new_vals):
                df.at[idx, c] = v
            df.at[idx, "_WasShifted"] = True
            fixed += 1
    return df, fixed


def fix_swapped_month_day(df, dt_cols):
    """Fix timestamps where SharePoint swapped month/day (e.g. 2026-10-06
    recorded for a scan on 2026-06-10). Heuristic per row: if a stage
    timestamp is > 30 days away from the row's median stamp AND swapping
    month/day brings it within 30 days, swap it."""
    fixed = 0
    for idx in df.index:
        stamps = {c: df.at[idx, c] for c in dt_cols
                  if c in df.columns and pd.notna(df.at[idx, c])}
        if len(stamps) < 2:
            continue
        med = pd.Series(list(stamps.values())).median()
        for c, ts in stamps.items():
            if abs((ts - med).days) > 30 and ts.month <= 12 and ts.day <= 12:
                try:
                    swapped = ts.replace(month=ts.day, day=ts.month)
                except ValueError:
                    continue
                if abs((swapped - med).days) <= 30:
                    df.at[idx, c] = swapped
                    fixed += 1
    return df, fixed


def load_scan_export(uploaded_file):
    """Read xlsx/csv into a DataFrame. Returns (df, error_message)."""
    name = uploaded_file.name.lower()
    if name.endswith(".iqy"):
        raw = uploaded_file.read().decode("utf-8", errors="ignore")
        url = next((ln.strip() for ln in raw.splitlines()
                    if ln.strip().lower().startswith("http")), None)
        msg = ("An **.iqy** file is not a data file — it is a *live web query* "
               "that tells Excel to pull the list from SharePoint using **your** "
               "login. This app has no way to authenticate to SharePoint, so it "
               "cannot fetch the data from it.\n\n"
               + (f"Query URL found inside: `{url}`\n\n" if url else "")
               + "**What to do instead:** open the .iqy in Excel (it will load "
                 "the live list), then *File → Save As →* `.xlsx` and upload "
                 "that here. Or in the SharePoint list use *Export → Export to "
                 "CSV*.")
        return None, msg
    try:
        if name.endswith(".csv"):
            df = pd.read_csv(uploaded_file)
        else:
            df = pd.read_excel(uploaded_file)
    except Exception as e:
        return None, f"Could not read the file: {e}"
    if "Title" not in df.columns:
        return None, "The file does not look like a ScanLog export (no 'Title' column)."
    return df, None


def find_duplicate_rows(df, dt_cols):
    """The ScanLog list can contain duplicate/test rows: different Title but
    timestamp-for-timestamp identical to a real insulator (no two physical
    units can share the same scan second across stages). Rule: a row is a
    duplicate if >=4 of its non-null timestamps exactly match a kept row.
    Lowest serial is kept. Returns (keep_df, dup_df)."""
    order = df.sort_values("Title").index.tolist()
    kept, dupes, kept_stamps = [], [], []
    for idx in order:
        stamps = {c: df.at[idx, c] for c in dt_cols
                  if c in df.columns and pd.notna(df.at[idx, c])}
        is_dup = None
        for kidx, kstamps in kept_stamps:
            matches = sum(1 for c, v in stamps.items()
                          if c in kstamps and kstamps[c] == v)
            if matches >= 4:
                is_dup = kidx
                break
        if is_dup is not None:
            dupes.append((idx, df.at[is_dup, "Title"]))
        else:
            kept.append(idx)
            kept_stamps.append((idx, stamps))
    dup_df = df.loc[[i for i, _ in dupes]].copy()
    if not dup_df.empty:
        dup_df["DuplicateOf"] = [t for _, t in dupes]
    return df.loc[kept].copy(), dup_df


def prepare_data(df, do_shift_repair=True, do_md_fix=True, do_dedup=True):
    """Clean + type the raw export. Returns (df, notes list, duplicates df)."""
    notes = []
    df = df.copy()
    # drop rows without an Insulator ID (blank SharePoint rows) + tidy key text cols
    blank = df["Title"].isna() | (df["Title"].astype(str).str.strip() == "")
    if blank.any():
        notes.append(f"Dropped {int(blank.sum())} blank row(s) with no Insulator ID.")
        df = df[~blank].copy()
    df["Title"] = df["Title"].astype(str).str.strip()
    for c in ("JobCard", "MRNo", "SKU"):
        if c in df.columns:
            df[c] = df[c].fillna("—").astype(str).str.strip().replace("", "—")
    all_dt_cols = [c for c in df.columns if c.endswith("_DT")]

    if do_shift_repair:
        df, n = repair_shifted_rows(df)
        if n:
            notes.append(f"Repaired {n} column-shifted row(s) (Rejection_MC held a timestamp).")

    for c in all_dt_cols:
        df[c] = pd.to_datetime(df[c], errors="coerce", utc=True).dt.tz_localize(None)

    if do_md_fix:
        df, n = fix_swapped_month_day(df, all_dt_cols)
        if n:
            notes.append(f"Corrected {n} swapped month/day timestamp(s).")

    dup_df = pd.DataFrame()
    if do_dedup:
        keep, clones = find_duplicate_rows(df, all_dt_cols)
        reasons = {}
        if not clones.empty:
            for _, r in clones.iterrows():
                reasons[r["Title"]] = f"Timestamp clone of {r['DuplicateOf']}"
        for _, r in df[df.get("_WasShifted", False) == True].iterrows():
            reasons.setdefault(r["Title"], "Column-shifted row (corrupted structure)")
        if reasons:
            dup_df = df[df["Title"].isin(reasons)].copy()
            dup_df["FlagReason"] = dup_df["Title"].map(reasons)
            df = df[~df["Title"].isin(reasons)].copy()
            notes.append(f"Quarantined {len(dup_df)} suspect row(s): timestamp "
                         "clones and/or structurally corrupted (column-shifted) "
                         "rows. Review the list — re-include any that are real.")

    # find rejection-reason column if the export carries one
    reason_col = next((c for c in df.columns if "reason" in c.lower()), None)
    df.attrs["reason_col"] = reason_col
    return df, notes, dup_df


def shift_window(d_from, d_to, shift_start):
    """Selected 6th→7th with 08:00 shift start = 6th 08:00 to 7th 08:00."""
    start = datetime.combine(d_from, shift_start)
    end = datetime.combine(d_to, shift_start)
    if end <= start:
        end = start + timedelta(days=1)
    return start, end


def detect_reused_rejected(df):
    """Rejected earlier, then rubber peeled and the SAME QR re-run from
    Priming: the process scans overwrite with NEW timestamps while the
    Rejection scan keeps the OLD one. Flag rows where Rejection_DT exists
    and a later production stage carries a NEWER timestamp."""
    reason_col = df.attrs.get("reason_col")
    out = []
    for _, r in df.iterrows():
        rej = r.get("Rejection_DT")
        if pd.isna(rej):
            continue
        later = {}
        for s in ["Priming", "Moulding", "CAM", "SML", "Cframe", "Packing", "Invoicing"]:
            ts = r.get(DT_COL[s])
            if pd.notna(ts) and ts > rej:
                later[s] = ts
        if later:
            first_stage = min(later, key=later.get)
            out.append({
                "InsulatorID": r["Title"], "JobCard": r.get("JobCard"),
                "MRNo": r.get("MRNo"), "SKU": r.get("SKU"),
                "Rejected At Stage": r.get("StageAsked") or "—",
                "Rejection Machine": r.get("Rejection_MC") or "—",
                "Rejection Reason": (r.get(reason_col) if reason_col else None) or "—",
                "Rejection DateTime": rej,
                "Reused From Stage": STAGE_LABELS.get(first_stage, first_stage),
                "Reuse DateTime": later[first_stage],
                "Latest Stage Reached": STAGE_LABELS.get(max(later, key=later.get)),
            })
    return pd.DataFrame(out)


DEFAULT_MACHINE_PREFIX = [("FP-145", "CBM"), ("FP-165", "CAM"),
                         ("SML", "SML"), ("M", "Moulding"), ("C", "Cframe")]


def infer_machine_map(df):
    """Machine -> Process mapping. Seeded from the scan file itself (a machine
    seen in CBM_MC belongs to CBM, etc.), then prefix rules for anything that
    only appears in Rework_MC / Rejection_MC."""
    mapping = {}
    for s, col in MC_COL.items():
        if col in df.columns:
            for m in df[col].dropna().astype(str).str.strip():
                if m and m != "—":
                    mapping.setdefault(m, s)
    extra = set()
    for col in ("Rework_MC", "Rejection_MC"):
        if col in df.columns:
            extra |= {str(m).strip() for m in df[col].dropna() if str(m).strip()}
    for m in sorted(extra - set(mapping)):
        for pref, s in DEFAULT_MACHINE_PREFIX:
            if m.upper().startswith(pref.upper()):
                mapping[m] = s
                break
    return mapping


def attribute_stage(stage_asked, machine, machine_map):
    """Which stage does a Rework/Rejection scan belong to?
    1) StageAsked picked by QC at scan time, 2) the machine's process
    from the mapping, 3) Unattributed."""
    if isinstance(stage_asked, str) and stage_asked.strip() in STAGES:
        return stage_asked.strip()
    m = (str(machine).strip() if pd.notna(machine) else "")
    if m and m in machine_map:
        return machine_map[m]
    return None


def classify_shift(ts, shift_a_start=time(8, 0)):
    """Shift A = shift_a_start -> +12h same day; Shift B = +12h -> shift_a_start
    next morning, credited to the STARTING date.
    E.g. 8th 21:30 -> (8th, 'B'); 9th 03:00 -> (8th, 'B'); 9th 09:00 -> (9th, 'A')."""
    a = shift_a_start
    b = time((a.hour + 12) % 24, a.minute)
    t = ts.time()
    if a <= t < b:
        return ts.date(), "A"
    if t >= b:
        return ts.date(), "B"
    return (ts - timedelta(days=1)).date(), "B"


def build_stage_frames(df, win_start, win_end, machine_map=None, shift_a_start=time(8, 0)):
    """UNIT-LEVEL stage tables (ERP semantics).

    Per stage: Produced = units whose stage scan falls in the window,
    plus units whose Rework/Rejection attributed to this stage falls in
    the window. Each unit appears ONCE per stage with a final Status
    (Rejection > Rework > OK) and OK Qty = Total - Rework - Rejection.
    Returns (frames, reused_in_window, reused_all, unattributed_df)."""
    machine_map = machine_map or infer_machine_map(df)
    reason_col = df.attrs.get("reason_col")
    reused = detect_reused_rejected(df)
    reused_ids = set(reused["InsulatorID"]) if not reused.empty else set()

    def in_win(ts):
        return pd.notna(ts) and win_start <= ts < win_end

    units = {s: {} for s in STAGES}     # stage -> {InsulatorID: record}
    unattributed = []

    def base_rec(r):
        return {"InsulatorID": r["Title"], "JobCard": r.get("JobCard"),
                "MRNo": r.get("MRNo"), "SKU": r.get("SKU"), "Status": "OK",
                "Prod Machine": "", "Prod DateTime": pd.NaT,
                "Rework Machine": "", "Rework DateTime": pd.NaT,
                "Rejection Machine": "", "Rejection DateTime": pd.NaT,
                "Rejection Reason": "", "Remarks": ""}

    for _, r in df.iterrows():
        # production scans
        for s in STAGES:
            ts = r.get(DT_COL[s])
            if in_win(ts):
                rec = units[s].setdefault(r["Title"], base_rec(r))
                mcc = MC_COL.get(s)
                rec["Prod Machine"] = (str(r.get(mcc)).strip()
                                       if mcc and pd.notna(r.get(mcc)) else "—")
                rec["Prod DateTime"] = ts
        # rework scan -> attributed stage
        ts = r.get("Rework_DT")
        if in_win(ts):
            s = attribute_stage(r.get("StageAsked"), r.get("Rework_MC"), machine_map)
            if s is None:
                unattributed.append({"InsulatorID": r["Title"], "Type": "Rework",
                                     "Machine": r.get("Rework_MC"),
                                     "StageAsked": r.get("StageAsked"),
                                     "DateTime": ts})
            else:
                rec = units[s].setdefault(r["Title"], base_rec(r))
                if rec["Status"] != "Rejection":
                    rec["Status"] = "Rework"
                rec["Rework Machine"] = (str(r.get("Rework_MC")).strip()
                                         if pd.notna(r.get("Rework_MC")) else "—")
                rec["Rework DateTime"] = ts
        # rejection scan -> attributed stage
        ts = r.get("Rejection_DT")
        if in_win(ts):
            s = attribute_stage(r.get("StageAsked"), r.get("Rejection_MC"), machine_map)
            if s is None:
                unattributed.append({"InsulatorID": r["Title"], "Type": "Rejection",
                                     "Machine": r.get("Rejection_MC"),
                                     "StageAsked": r.get("StageAsked"),
                                     "DateTime": ts})
            else:
                rec = units[s].setdefault(r["Title"], base_rec(r))
                rec["Status"] = "Rejection"
                rec["Rejection Machine"] = (str(r.get("Rejection_MC")).strip()
                                            if pd.notna(r.get("Rejection_MC")) else "—")
                rec["Rejection DateTime"] = ts
                rec["Rejection Reason"] = ((r.get(reason_col) if reason_col else None)
                                           or "")
                if r["Title"] in reused_ids:
                    rec["Remarks"] = "REUSED LATER — QR re-run from Priming"

    result = {}
    for s in STAGES:
        det = pd.DataFrame(list(units[s].values()))
        if det.empty:
            result[s] = (pd.DataFrame(), det)
            continue
        det = det.sort_values(["JobCard", "MRNo", "InsulatorID"]).reset_index(drop=True)
        # event timestamp that defines WHEN the unit's stage outcome happened
        det["EventDT"] = det.apply(
            lambda r: r["Rejection DateTime"] if r["Status"] == "Rejection"
            else (r["Rework DateTime"] if r["Status"] == "Rework"
                  else r["Prod DateTime"]), axis=1)
        sd = det["EventDT"].apply(
            lambda t: classify_shift(t, shift_a_start) if pd.notna(t) else (None, ""))
        det["Shift Date"] = sd.apply(
            lambda x: x[0].strftime("%d/%m/%Y") if x[0] else "")
        det["Shift"] = sd.apply(lambda x: x[1])
        grp = det.groupby(["JobCard", "MRNo"])
        summ = pd.DataFrame({
            "Total Produced Qty": grp.size(),
            "Rework Qty": grp["Status"].apply(lambda g: int((g == "Rework").sum())),
            "Rejection Qty": grp["Status"].apply(lambda g: int((g == "Rejection").sum())),
        }).reset_index().rename(columns={"JobCard": "Job Card", "MRNo": "MR No"})
        summ["OK Qty"] = (summ["Total Produced Qty"] - summ["Rework Qty"]
                          - summ["Rejection Qty"])
        summ = summ[["Job Card", "MR No", "Total Produced Qty", "OK Qty",
                     "Rework Qty", "Rejection Qty"]]
        result[s] = (summ, det)

    if not reused.empty:
        m = reused.apply(lambda r: (win_start <= r["Reuse DateTime"] < win_end) or
                                   (win_start <= r["Rejection DateTime"] < win_end),
                         axis=1)
        reused_win = reused[m].reset_index(drop=True)
    else:
        reused_win = reused
    return result, reused_win, reused, pd.DataFrame(unattributed)


def unit_level_status(stage_frames):
    """One final status per distinct insulator across all stages in the window.
    Rejection anywhere > Rework anywhere > OK (clean through every stage)."""
    rank = {"OK": 0, "Rework": 1, "Rejection": 2}
    status = {}
    for _, det in stage_frames.values():
        if det.empty:
            continue
        for _, r in det.iterrows():
            iid, s = r["InsulatorID"], r["Status"]
            if iid not in status or rank[s] > rank[status[iid]]:
                status[iid] = s
    return status


# ======================================================================
# Excel report writer (openpyxl, formatted, COUNTIFS formulas)
# ======================================================================
def write_excel_report(stage_frames, reused_win, win_start, win_end, shift_a_start=time(8, 0)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    hdr_fill = PatternFill("solid", start_color="1B5EA6")
    sub_fill = PatternFill("solid", start_color="E8F0F9")
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    title_font = Font(name="Arial", bold=True, color="134A85", size=13)
    sub_font = Font(name="Arial", bold=True, color="134A85", size=11)
    body_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="B9CDE4")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    wb = Workbook()
    wb.remove(wb.active)

    def style_header(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = center; cell.border = box

    def sheet_title(ws, text, sub):
        ws["A1"] = text; ws["A1"].font = title_font
        ws["A2"] = sub; ws["A2"].font = Font(name="Arial", size=9, color="6B7280")

    win_txt = f"Shift window: {win_start:%d/%m/%Y %H:%M} to {win_end:%d/%m/%Y %H:%M}"

    # ---- Overview sheet: Job Card × Stage breakdown ----------------------
    ov = wb.create_sheet("Overview")
    sheet_title(ov, "Deccan Enterprises — Daily Production Entry Report", win_txt)

    # collect all job cards present in the window, per stage
    all_jcs = sorted({jc for s in STAGES for jc in
                      (stage_frames[s][0]["Job Card"].tolist()
                       if not stage_frames[s][0].empty else [])})
    ov["A4"] = "Job Card"; ov["B4"] = "Stage"; ov["C4"] = "Total Produced"
    ov["D4"] = "OK"; ov["E4"] = "Rework"; ov["F4"] = "Rejection"
    style_header(ov, 4, 6)
    r = 5
    for jc in all_jcs:
        block_start = r
        for s in STAGES:
            summ, det = stage_frames[s]
            if det.empty or jc not in set(det["JobCard"]):
                continue
            sn = f"'{STAGE_LABELS[s]}'"
            ov.cell(row=r, column=1, value=jc).font = body_font
            ov.cell(row=r, column=2, value=STAGE_LABELS[s]).font = body_font
            ov.cell(row=r, column=3, value=f"=SUM(D{r}:F{r})")
            ov.cell(row=r, column=4,
                    value=f'=COUNTIFS({sn}!$B:$B,$A{r},{sn}!$E:$E,"OK")')
            ov.cell(row=r, column=5,
                    value=f'=COUNTIFS({sn}!$B:$B,$A{r},{sn}!$E:$E,"Rework")')
            ov.cell(row=r, column=6,
                    value=f'=COUNTIFS({sn}!$B:$B,$A{r},{sn}!$E:$E,"Rejection")')
            for c in range(1, 7):
                ov.cell(row=r, column=c).border = box
                if c > 2:
                    ov.cell(row=r, column=c).alignment = center
                    ov.cell(row=r, column=c).font = body_font
            r += 1
        if r > block_start:
            # ---- job card units summary (a unit passing N stages is still 1 unit) ----
            units_total = len({iid for s in STAGES
                               for iid in (stage_frames[s][1]["InsulatorID"]
                                           [stage_frames[s][1]["JobCard"] == jc].tolist()
                                           if not stage_frames[s][1].empty else [])})
            row_units, row_rej, row_net = r, r + 1, r + 2
            ov.cell(row=row_units, column=1, value=jc).font = sub_font
            ov.cell(row=row_units, column=2,
                    value="TOTAL UNITS IN PRODUCTION").font = sub_font
            ov.cell(row=row_units, column=3, value=units_total).font = sub_font
            ov.cell(row=row_rej, column=1, value=jc).font = sub_font
            ov.cell(row=row_rej, column=2, value="REJECTED UNITS").font = sub_font
            ov.cell(row=row_rej, column=3,
                    value=f"=SUM(F{block_start}:F{r-1})").font = sub_font
            ov.cell(row=row_net, column=1, value=jc).font = sub_font
            ov.cell(row=row_net, column=2,
                    value="NET PRODUCED (Units − Rejected)").font = sub_font
            ov.cell(row=row_net, column=3,
                    value=f"=C{row_units}-C{row_rej}").font = sub_font
            for rw in (row_units, row_rej, row_net):
                for c in range(1, 7):
                    ov.cell(row=rw, column=c).border = box
                    ov.cell(row=rw, column=c).fill = sub_fill
                ov.cell(row=rw, column=3).alignment = center
            r = row_net + 2                          # blank spacer between job cards
    if not all_jcs:
        ov.cell(row=r, column=1,
                value="No records in the selected shift window.").font = body_font
        r += 1
    ov.cell(row=r, column=1, value="Reused Rejected (flagged)").font = sub_font
    ov.cell(row=r, column=2, value=len(reused_win)).font = sub_font
    for w_, col in zip([16, 20, 16, 10, 10, 12], "ABCDEF"):
        ov.column_dimensions[col].width = w_
    ov.freeze_panes = "A5"

    # ---- Stage sheets ---------------------------------------------------
    for s in STAGES:
        summ, det = stage_frames[s]
        n_shift_rows = (det[det["Shift Date"] != ""]
                        .groupby(["Shift Date", "Shift", "JobCard", "MRNo"])
                        .ngroups if not det.empty else 0)
        DETAIL_START = max(100, len(summ) + n_shift_rows + 22)
        ws = wb.create_sheet(STAGE_LABELS[s][:31])
        sheet_title(ws, f"{STAGE_LABELS[s]} — Production Entry", win_txt)

        ws["A4"] = "SUMMARY (Job Card + MR level — enter these in ERP)"
        ws["A4"].font = sub_font; ws["A4"].fill = sub_fill
        heads = ["Job Card", "MR No", "Total Produced Qty", "OK Qty",
                 "Rework Qty", "Rejection Qty"]
        for c, h in enumerate(heads, 1):
            ws.cell(row=5, column=c, value=h)
        style_header(ws, 5, len(heads))

        rr = 6
        if not summ.empty:
            D = DETAIL_START
            for _, row in summ.iterrows():
                ws.cell(row=rr, column=1, value=row["Job Card"]).font = body_font
                ws.cell(row=rr, column=2, value=row["MR No"]).font = body_font
                # live COUNTIFS against the detail table below (JobCard=B, MR=C, Status=E)
                ws.cell(row=rr, column=3,
                        value=f'=COUNTIFS($B${D}:$B$100000,A{rr},$C${D}:$C$100000,B{rr})')
                ws.cell(row=rr, column=4,
                        value=f'=COUNTIFS($B${D}:$B$100000,A{rr},$C${D}:$C$100000,B{rr},$E${D}:$E$100000,"OK")')
                ws.cell(row=rr, column=5,
                        value=f'=COUNTIFS($B${D}:$B$100000,A{rr},$C${D}:$C$100000,B{rr},$E${D}:$E$100000,"Rework")')
                ws.cell(row=rr, column=6,
                        value=f'=COUNTIFS($B${D}:$B$100000,A{rr},$C${D}:$C$100000,B{rr},$E${D}:$E$100000,"Rejection")')
                for c in range(1, 7):
                    ws.cell(row=rr, column=c).border = box
                    if c > 2:
                        ws.cell(row=rr, column=c).alignment = center
                        ws.cell(row=rr, column=c).font = body_font
                rr += 1
            ws.cell(row=rr, column=1, value="TOTAL").font = sub_font
            for c in range(3, 7):
                col = get_column_letter(c)
                cell = ws.cell(row=rr, column=c, value=f"=SUM({col}6:{col}{rr-1})")
                cell.font = sub_font; cell.alignment = center; cell.border = box
            ws.cell(row=rr, column=1).border = box
            ws.cell(row=rr, column=2).border = box
        else:
            ws.cell(row=rr, column=1, value="No records in the selected shift window.").font = body_font

        # ---- SHIFT-WISE SUMMARY (Date x Shift, merged date cells) ----
        if not det.empty:
            b_start = time((shift_a_start.hour + 12) % 24, shift_a_start.minute)
            a_txt = (f"Shift A  {shift_a_start:%H:%M}–{b_start:%H:%M}")
            b_txt = (f"Shift B  {b_start:%H:%M}–{shift_a_start:%H:%M} (+1 day)")
            srow = rr + 3
            hdr = ws.cell(row=srow, column=1,
                          value="SHIFT-WISE SUMMARY (Date × Shift)")
            hdr.font = sub_font; hdr.fill = sub_fill
            srow += 1
            sheads = ["Date", "Shift", "Job Card", "MR No",
                      "Total Produced Qty", "OK Qty", "Rework Qty", "Rejection Qty"]
            for c, h in enumerate(sheads, 1):
                ws.cell(row=srow, column=c, value=h)
            style_header(ws, srow, len(sheads))
            srow += 1
            D = DETAIL_START
            grouped = (det[det["Shift Date"] != ""]
                       .groupby(["Shift Date", "Shift", "JobCard", "MRNo"])
                       .size().reset_index(name="_n"))
            grouped["_d"] = pd.to_datetime(grouped["Shift Date"], format="%d/%m/%Y")
            grouped = grouped.sort_values(["_d", "Shift", "JobCard", "MRNo"])
            for dt_txt, ddf in grouped.groupby("Shift Date", sort=False):
                d0 = srow
                for _, g in ddf.iterrows():
                    sh = g["Shift"]
                    ws.cell(row=srow, column=2,
                            value=a_txt if sh == "A" else b_txt).font = body_font
                    ws.cell(row=srow, column=3, value=g["JobCard"]).font = body_font
                    ws.cell(row=srow, column=4, value=g["MRNo"]).font = body_font
                    base = (f'$B${D}:$B$100000,C{srow},$C${D}:$C$100000,D{srow},'
                            f'$N${D}:$N$100000,A{d0},$O${D}:$O$100000,"{sh}"')
                    ws.cell(row=srow, column=5, value=f'=COUNTIFS({base})')
                    ws.cell(row=srow, column=6,
                            value=f'=COUNTIFS({base},$E${D}:$E$100000,"OK")')
                    ws.cell(row=srow, column=7,
                            value=f'=COUNTIFS({base},$E${D}:$E$100000,"Rework")')
                    ws.cell(row=srow, column=8,
                            value=f'=COUNTIFS({base},$E${D}:$E$100000,"Rejection")')
                    for c in range(1, 9):
                        ws.cell(row=srow, column=c).border = box
                        if c >= 5:
                            ws.cell(row=srow, column=c).alignment = center
                            ws.cell(row=srow, column=c).font = body_font
                    srow += 1
                # merged date cell spanning this date's shift rows
                ws.cell(row=d0, column=1, value=dt_txt)
                if srow - 1 > d0:
                    ws.merge_cells(start_row=d0, start_column=1,
                                   end_row=srow - 1, end_column=1)
                mc = ws.cell(row=d0, column=1)
                mc.font = sub_font; mc.alignment = Alignment(
                    horizontal="center", vertical="center"); mc.fill = sub_fill

        ws.cell(row=DETAIL_START - 2, column=1,
                value="DETAIL (Insulator level)").font = sub_font
        ws.cell(row=DETAIL_START - 2, column=1).fill = sub_fill
        dheads = ["Insulator ID", "Job Card", "MR No", "SKU", "Status",
                  "Prod Machine", "Prod Date", "Prod Time",
                  "Rework Machine", "Rework Date/Time",
                  "Rejection Machine", "Rejection Date/Time",
                  "Rejection Reason / Remarks", "Shift Date", "Shift"]
        for c, h in enumerate(dheads, 1):
            ws.cell(row=DETAIL_START - 1, column=c, value=h)
        style_header(ws, DETAIL_START - 1, len(dheads))

        def fmt_dt(ts, f="%d/%m/%Y %H:%M"):
            return ts.strftime(f) if pd.notna(ts) else ""

        rr = DETAIL_START
        if not det.empty:
            for jc, grp_df in det.groupby("JobCard", sort=True):
                band = ws.cell(row=rr, column=1, value=f"Job Card: {jc}")
                band.font = sub_font
                for c in range(1, len(dheads) + 1):
                    ws.cell(row=rr, column=c).fill = sub_fill
                    ws.cell(row=rr, column=c).border = box
                rr += 1
                for _, row in grp_df.iterrows():
                    reason = row["Rejection Reason"]
                    remarks = row["Remarks"]
                    reason_txt = " — ".join(x for x in [reason, remarks] if x)
                    vals = [row["InsulatorID"], row["JobCard"], row["MRNo"],
                            row["SKU"], row["Status"], row["Prod Machine"] or "—",
                            fmt_dt(row["Prod DateTime"], "%d/%m/%Y"),
                            fmt_dt(row["Prod DateTime"], "%H:%M:%S"),
                            row["Rework Machine"] or "",
                            fmt_dt(row["Rework DateTime"]),
                            row["Rejection Machine"] or "",
                            fmt_dt(row["Rejection DateTime"]),
                            reason_txt, row["Shift Date"], row["Shift"]]
                    for c, v in enumerate(vals, 1):
                        cell = ws.cell(row=rr, column=c, value=v)
                        cell.font = body_font; cell.border = box
                        if c in (5, 6, 7, 8, 9, 10, 11, 12):
                            cell.alignment = center
                    st_ = row["Status"]
                    if st_ == "Rejection":
                        ws.cell(row=rr, column=5).font = Font(name="Arial", size=10,
                                                              bold=True, color="C0392B")
                    elif st_ == "Rework":
                        ws.cell(row=rr, column=5).font = Font(name="Arial", size=10,
                                                              bold=True, color="B9770E")
                    rr += 1
                rr += 1

        # ---- REWORK & REJECTION DETAILS section ----
        rrej = det[det["Status"].isin(["Rework", "Rejection"])] if not det.empty \
            else pd.DataFrame()
        if not rrej.empty:
            rr += 1
            hdr = ws.cell(row=rr, column=1,
                          value="REWORK & REJECTION DETAILS — reasons and machines")
            hdr.font = sub_font; hdr.fill = sub_fill
            rr += 1
            rheads = ["Insulator ID", "Type", "Machine", "Stage", "Job Card",
                      "MR No", "Date/Time", "Reason / Remarks"]
            for c, h in enumerate(rheads, 1):
                ws.cell(row=rr, column=c, value=h)
            style_header(ws, rr, len(rheads))
            rr += 1
            for _, row in rrej.sort_values(["Status", "InsulatorID"]).iterrows():
                is_rej = row["Status"] == "Rejection"
                machine = (row["Rejection Machine"] if is_rej
                           else row["Rework Machine"]) or "—"
                ts = (row["Rejection DateTime"] if is_rej
                      else row["Rework DateTime"])
                reason = " — ".join(x for x in
                                    [row["Rejection Reason"] if is_rej else "",
                                     row["Remarks"]] if x) or "—"
                vals = [row["InsulatorID"], row["Status"], machine,
                        STAGE_LABELS[s], row["JobCard"], row["MRNo"],
                        fmt_dt(ts), reason]
                for c, v in enumerate(vals, 1):
                    cell = ws.cell(row=rr, column=c, value=v)
                    cell.font = body_font; cell.border = box
                    if c in (2, 3, 4, 7):
                        cell.alignment = center
                ws.cell(row=rr, column=2).font = Font(
                    name="Arial", size=10, bold=True,
                    color="C0392B" if is_rej else "B9770E")
                rr += 1
        for w_, col in zip([22, 13, 15, 11, 11, 12, 11, 9, 13, 16, 15, 17, 34, 12, 7],
                           ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O"]):
            ws.column_dimensions[col].width = w_
        ws.freeze_panes = "A6"

    # ---- Reused Rejected sheet ------------------------------------------
    ws = wb.create_sheet("Reused Rejected")
    sheet_title(ws, "Reused Rejected Quantity — rejected units re-run from Priming (same QR)",
                win_txt)
    ws["A3"] = ("These units carry an OLD Rejection timestamp but NEWER production "
                "timestamps — the rubber was peeled and the rod restarted from Priming.")
    ws["A3"].font = Font(name="Arial", size=9, italic=True, color="6B7280")
    heads = ["Insulator ID", "Job Card", "MR No", "SKU", "Rejected At Stage",
             "Rejection Machine", "Rejection Reason", "Rejection Date/Time",
             "Reused From Stage", "Reuse Date/Time", "Latest Stage Reached"]
    for c, h in enumerate(heads, 1):
        ws.cell(row=5, column=c, value=h)
    style_header(ws, 5, len(heads))
    rr = 6
    if not reused_win.empty:
        for _, row in reused_win.iterrows():
            vals = [row["InsulatorID"], row["JobCard"], row["MRNo"], row["SKU"],
                    row["Rejected At Stage"], row["Rejection Machine"],
                    row["Rejection Reason"],
                    row["Rejection DateTime"].strftime("%d/%m/%Y %H:%M"),
                    row["Reused From Stage"],
                    row["Reuse DateTime"].strftime("%d/%m/%Y %H:%M"),
                    row["Latest Stage Reached"]]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=rr, column=c, value=v)
                cell.font = body_font; cell.border = box
            rr += 1
        ws.cell(row=rr + 1, column=1, value="Reused Rejected Quantity:").font = sub_font
        ws.cell(row=rr + 1, column=2, value=f"=COUNTA(A6:A{rr-1})").font = sub_font
    else:
        ws.cell(row=6, column=1,
                value="No reused-rejected units detected in the selected window.").font = body_font
    for w_, col in zip([22, 14, 16, 12, 16, 16, 22, 18, 16, 18, 18],
                       ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]):
        ws.column_dimensions[col].width = w_

    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


# ======================================================================
# Charts (Altair — bundled with Streamlit, no extra install needed)
# ======================================================================
def render_charts(stage_frames, reused_all):
    import altair as alt

    STATUS_ORDER = ["OK", "Rework", "Rejection"]
    STATUS_COLORS = ["#2E8B57", "#E8A317", "#C0392B"]
    status_scale = alt.Scale(domain=STATUS_ORDER, range=STATUS_COLORS)
    rank = {"OK": 0, "Rework": 1, "Rejection": 2}
    GRID = {"grid": True, "gridColor": "#dbe6f2", "labelFontSize": 12,
            "titleFontSize": 13, "titleColor": "#134A85"}

    all_det = pd.concat(
        [d.assign(Stage=STAGE_LABELS[s]) for s, (m, d) in stage_frames.items()
         if not d.empty], ignore_index=True) \
        if any(not d.empty for _, d in stage_frames.values()) else pd.DataFrame()
    if all_det.empty:
        st.info("No scan records in the selected window — nothing to chart.")
        return

    st.markdown("#### Filters")
    f1, f2, f3 = st.columns(3)
    jcs = sorted(all_det["JobCard"].dropna().unique())
    stages_avail = [STAGE_LABELS[s] for s in STAGES if not stage_frames[s][1].empty]
    sel_jc = f1.multiselect("Job Card", jcs, default=jcs)
    sel_st = f2.multiselect("Stage", stages_avail, default=stages_avail)
    sel_status = f3.multiselect("Status", STATUS_ORDER, default=STATUS_ORDER)
    fdf = all_det[all_det["JobCard"].isin(sel_jc) & all_det["Stage"].isin(sel_st)
                  & all_det["Status"].isin(sel_status)].copy()
    if fdf.empty:
        st.warning("No data after filters."); return

    charts_for_export = []

    def register(chart):
        charts_for_export.append(chart)
        return chart

    # ---------- 1. Per-stage quality donuts (each stage: OK / Rework / Rejection) ----------
    st.markdown("##### Stage-wise Quality Split")
    st.caption("One donut per stage — units that passed that stage and their "
               "status AT that stage. Numbers on the slices are unit counts.")
    stage_order = [STAGE_LABELS[s] for s in STAGES]
    donuts = []
    for s_lbl in [x for x in stage_order if x in set(fdf["Stage"])]:
        sdf = fdf[fdf["Stage"] == s_lbl]
        g = (sdf["Status"].value_counts().rename_axis("Status")
             .reset_index(name="Qty"))
        total = int(g["Qty"].sum())
        base = alt.Chart(g).encode(
            theta=alt.Theta("Qty:Q", stack=True),
            color=alt.Color("Status:N", scale=status_scale,
                            legend=alt.Legend(title=None, orient="bottom")),
            tooltip=["Status", "Qty"])
        arc = base.mark_arc(innerRadius=48, outerRadius=78)
        labels = base.mark_text(radius=95, fontWeight="bold",
                                fontSize=13).encode(text="Qty:Q")
        centre = alt.Chart(pd.DataFrame({"t": [f"{total}"]})).mark_text(
            fontSize=20, fontWeight="bold", color="#134A85").encode(text="t:N")
        donuts.append((arc + labels + centre).properties(
            title=alt.TitleParams(f"{s_lbl} — {total} units", fontSize=13,
                                  color="#134A85"),
            width=190, height=210))
    for i in range(0, len(donuts), 4):
        row = alt.hconcat(*donuts[i:i + 4])
        register(row)
        st.altair_chart(row, width="stretch")

    # ---------- 2. Stage-wise stacked bar ----------
    g = (fdf.groupby(["Stage", "Status"]).size().reset_index(name="Qty"))
    base = alt.Chart(g).encode(
        x=alt.X("Stage:N", sort=stage_order, title="Production Stage",
                axis=alt.Axis(labelAngle=0, **{k: v for k, v in GRID.items()
                                               if k != "grid"})),
        y=alt.Y("Qty:Q", title="Units", axis=alt.Axis(**GRID),
                scale=alt.Scale(nice=True)),
        tooltip=["Stage", "Status", "Qty"])
    bars = base.mark_bar().encode(color=alt.Color("Status:N", scale=status_scale,
                                                  legend=alt.Legend(title="Status")))
    text = base.mark_text(dy=-6, fontWeight="bold").encode(
        text="Qty:Q", y=alt.Y("Qty:Q", stack="zero"))
    ch = (bars + text).properties(title="Units at each Stage (OK / Rework / Rejection)",
                                  height=340)
    register(ch)
    c1, c2 = st.columns(2)
    with c1:
        st.altair_chart(ch, width="stretch")

    # ---------- 3. Job Card — unit-level (a unit is counted ONCE) ----------
    ustat = (fdf.groupby(["JobCard", "InsulatorID"])["Status"]
             .agg(lambda g_: max(g_, key=lambda s_: rank[s_])).reset_index())
    g = ustat.groupby(["JobCard", "Status"]).size().reset_index(name="Qty")
    base = alt.Chart(g).encode(
        x=alt.X("JobCard:N", title="Job Card",
                axis=alt.Axis(labelAngle=0, labelFontSize=12, titleFontSize=13,
                              titleColor="#134A85")),
        y=alt.Y("Qty:Q", title="Distinct Insulators", axis=alt.Axis(**GRID)),
        xOffset="Status:N", tooltip=["JobCard", "Status", "Qty"])
    bars = base.mark_bar().encode(color=alt.Color("Status:N", scale=status_scale,
                                                  legend=alt.Legend(title="Final Status")))
    text = base.mark_text(dy=-6, fontWeight="bold").encode(text="Qty:Q")
    ch = (bars + text).properties(
        title="Job Card — Distinct Units by Final Status (counted once)", height=340)
    register(ch)
    with c2:
        st.altair_chart(ch, width="stretch")

    # ---------- 4. Machine-wise + hourly throughput ----------
    c3, c4 = st.columns(2)
    with c3:
        mdf = fdf.assign(Machine=fdf["Prod Machine"].where(
            fdf["Status"] == "OK",
            fdf["Rework Machine"].where(fdf["Status"] == "Rework",
                                        fdf["Rejection Machine"])))
        mdf = mdf[mdf["Machine"].astype(str).str.strip().ne("") &
                  mdf["Machine"].ne("—")]
        if not mdf.empty:
            g = mdf.groupby(["Machine", "Status"]).size().reset_index(name="Qty")
            base = alt.Chart(g).encode(
                x=alt.X("Machine:N", title="Machine",
                        axis=alt.Axis(labelAngle=0, labelFontSize=12,
                                      titleFontSize=13, titleColor="#134A85")),
                y=alt.Y("Qty:Q", title="Units", axis=alt.Axis(**GRID)),
                tooltip=["Machine", "Status", "Qty"])
            bars = base.mark_bar().encode(
                color=alt.Color("Status:N", scale=status_scale,
                                legend=alt.Legend(title="Status")))
            text = base.mark_text(dy=-6, fontWeight="bold").encode(
                text="Qty:Q", y=alt.Y("Qty:Q", stack="zero"))
            ch = (bars + text).properties(title="Machine-wise Output", height=340)
            register(ch)
            st.altair_chart(ch, width="stretch")
    with c4:
        okd = fdf[pd.notna(fdf["Prod DateTime"])].copy()
        if not okd.empty:
            okd["Hour"] = okd["Prod DateTime"].dt.floor("h")
            g = okd.groupby(["Hour", "Stage"]).size().reset_index(name="Qty")
            ch = alt.Chart(g).mark_line(point=alt.OverlayMarkDef(size=70)).encode(
                x=alt.X("Hour:T", title="Date / Hour",
                        axis=alt.Axis(format="%d/%m %H:%M", labelAngle=-30, **GRID)),
                y=alt.Y("Qty:Q", title="Units scanned", axis=alt.Axis(**GRID)),
                color=alt.Color("Stage:N", legend=alt.Legend(title="Stage"),
                                sort=stage_order),
                tooltip=[alt.Tooltip("Hour:T", format="%d/%m/%Y %H:%M"),
                         "Stage", "Qty"]).properties(
                title="Hourly Throughput by Stage", height=340)
            register(ch)
            st.altair_chart(ch, width="stretch")

    # ---------- 5. Reused rejected ----------
    if not reused_all.empty:
        g = reused_all.groupby("Rejected At Stage").size().reset_index(name="Qty")
        base = alt.Chart(g).encode(
            x=alt.X("Rejected At Stage:N", title="Stage where rejection happened",
                    axis=alt.Axis(labelAngle=0, labelFontSize=12, titleFontSize=13,
                                  titleColor="#134A85")),
            y=alt.Y("Qty:Q", title="Units", axis=alt.Axis(**GRID)),
            tooltip=["Rejected At Stage", "Qty"])
        ch = (base.mark_bar(color=DECCAN_BLUE)
              + base.mark_text(dy=-6, fontWeight="bold").encode(text="Qty:Q")
              ).properties(title="Reused Rejected Units by Rejection Stage",
                           height=300)
        register(ch)
        st.altair_chart(ch, width="stretch")

    # ---------- Export: ALL charts as one PNG or PDF (one click) ----------
    st.markdown("##### Export all charts")
    st.caption("One click composes every chart above into a single image — "
               "download as PNG or PDF.")
    import json as _json
    import streamlit.components.v1 as components
    specs = _json.dumps([c.to_dict() for c in charts_for_export])
    ts_txt = datetime.now().strftime("%d/%m/%Y %H:%M")
    components.html(f"""
    <div style="font-family:'Source Sans Pro',Arial,sans-serif">
      <button id="btnPng" style="background:#1B5EA6;color:#fff;border:none;
              border-radius:6px;padding:10px 18px;font-weight:600;cursor:pointer;
              margin-right:10px">⬇ Download ALL charts (PNG)</button>
      <button id="btnPdf" style="background:#134A85;color:#fff;border:none;
              border-radius:6px;padding:10px 18px;font-weight:600;cursor:pointer">
              ⬇ Download ALL charts (PDF)</button>
      <span id="stat" style="margin-left:12px;color:#6B7280"></span>
      <div id="work" style="position:absolute;left:-99999px;top:0"></div>
    </div>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/vega/5.30.0/vega.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/vega-lite/5.21.0/vega-lite.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/vega-embed/6.26.0/vega-embed.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/jspdf/2.5.1/jspdf.umd.min.js"></script>
    <script>
    const SPECS = {specs};
    const TITLE = "Deccan Enterprises — Production Analytics";
    const SUB = "Generated {ts_txt}";
    const stat = document.getElementById('stat');

    async function renderAll() {{
      stat.textContent = 'Rendering charts…';
      const work = document.getElementById('work');
      work.innerHTML = '';
      const urls = [];
      for (let i = 0; i < SPECS.length; i++) {{
        const d = document.createElement('div');
        work.appendChild(d);
        const res = await vegaEmbed(d, SPECS[i], {{actions: false}});
        urls.push(await res.view.toImageURL('png', 2));
        stat.textContent = `Rendering charts… ${{i+1}}/${{SPECS.length}}`;
      }}
      const imgs = await Promise.all(urls.map(u => new Promise((ok, err) => {{
        const im = new Image(); im.onload = () => ok(im); im.onerror = err; im.src = u;
      }})));
      const PAD = 40, HEAD = 110;
      const w = Math.max(...imgs.map(i => i.width)) + PAD * 2;
      const h = HEAD + imgs.reduce((a, i) => a + i.height + PAD, 0) + PAD;
      const cv = document.createElement('canvas');
      cv.width = w; cv.height = h;
      const ctx = cv.getContext('2d');
      ctx.fillStyle = '#FFFFFF'; ctx.fillRect(0, 0, w, h);
      ctx.fillStyle = '#134A85';
      ctx.font = 'bold 40px Arial'; ctx.textAlign = 'center';
      ctx.fillText(TITLE, w / 2, 55);
      ctx.fillStyle = '#6B7280'; ctx.font = '22px Arial';
      ctx.fillText(SUB, w / 2, 90);
      let y = HEAD;
      for (const im of imgs) {{
        ctx.drawImage(im, (w - im.width) / 2, y);
        y += im.height + PAD;
      }}
      work.innerHTML = '';
      return cv;
    }}
    document.getElementById('btnPng').onclick = async () => {{
      try {{
        const cv = await renderAll();
        const a = document.createElement('a');
        a.download = 'Deccan_Production_Charts.png';
        a.href = cv.toDataURL('image/png');
        a.click();
        stat.textContent = 'PNG downloaded ✓';
      }} catch (e) {{ stat.textContent = 'Failed: ' + e; }}
    }};
    document.getElementById('btnPdf').onclick = async () => {{
      try {{
        const cv = await renderAll();
        const pdf = new jspdf.jsPDF({{orientation: cv.width > cv.height ? 'l' : 'p',
                                     unit: 'px', format: [cv.width, cv.height],
                                     compress: true}});
        pdf.addImage(cv.toDataURL('image/jpeg', 0.92), 'JPEG', 0, 0,
                     cv.width, cv.height);
        pdf.save('Deccan_Production_Charts.pdf');
        stat.textContent = 'PDF downloaded ✓';
      }} catch (e) {{ stat.textContent = 'Failed: ' + e; }}
    }};
    </script>
    """, height=70)


def unit_level_status(stage_frames):
    """One final status per distinct insulator across all stages in the window.
    Rejection anywhere > Rework anywhere > OK (clean through every stage)."""
    rank = {"OK": 0, "Rework": 1, "Rejection": 2}
    status = {}
    for _, det in stage_frames.values():
        if det.empty:
            continue
        for _, r in det.iterrows():
            iid, s = r["InsulatorID"], r["Status"]
            if iid not in status or rank[s] > rank[status[iid]]:
                status[iid] = s
    return status


# ======================================================================
# Excel report writer (openpyxl, formatted, COUNTIFS formulas)
# ======================================================================
def write_excel_report(stage_frames, reused_win, win_start, win_end, shift_a_start=time(8, 0)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    hdr_fill = PatternFill("solid", start_color="1B5EA6")
    sub_fill = PatternFill("solid", start_color="E8F0F9")
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    title_font = Font(name="Arial", bold=True, color="134A85", size=13)
    sub_font = Font(name="Arial", bold=True, color="134A85", size=11)
    body_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="B9CDE4")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    wb = Workbook()
    wb.remove(wb.active)

    def style_header(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = center; cell.border = box

    def sheet_title(ws, text, sub):
        ws["A1"] = text; ws["A1"].font = title_font
        ws["A2"] = sub; ws["A2"].font = Font(name="Arial", size=9, color="6B7280")

    win_txt = f"Shift window: {win_start:%d/%m/%Y %H:%M} to {win_end:%d/%m/%Y %H:%M}"

    # ---- Overview sheet: Job Card × Stage breakdown ----------------------
    ov = wb.create_sheet("Overview")
    sheet_title(ov, "Deccan Enterprises — Daily Production Entry Report", win_txt)

    # collect all job cards present in the window, per stage
    all_jcs = sorted({jc for s in STAGES for jc in
                      (stage_frames[s][0]["Job Card"].tolist()
                       if not stage_frames[s][0].empty else [])})
    ov["A4"] = "Job Card"; ov["B4"] = "Stage"; ov["C4"] = "Total Produced"
    ov["D4"] = "OK"; ov["E4"] = "Rework"; ov["F4"] = "Rejection"
    style_header(ov, 4, 6)
    r = 5
    for jc in all_jcs:
        block_start = r
        for s in STAGES:
            summ, det = stage_frames[s]
            if det.empty or jc not in set(det["JobCard"]):
                continue
            sn = f"'{STAGE_LABELS[s]}'"
            ov.cell(row=r, column=1, value=jc).font = body_font
            ov.cell(row=r, column=2, value=STAGE_LABELS[s]).font = body_font
            ov.cell(row=r, column=3, value=f"=SUM(D{r}:F{r})")
            ov.cell(row=r, column=4,
                    value=f'=COUNTIFS({sn}!$B:$B,$A{r},{sn}!$E:$E,"OK")')
            ov.cell(row=r, column=5,
                    value=f'=COUNTIFS({sn}!$B:$B,$A{r},{sn}!$E:$E,"Rework")')
            ov.cell(row=r, column=6,
                    value=f'=COUNTIFS({sn}!$B:$B,$A{r},{sn}!$E:$E,"Rejection")')
            for c in range(1, 7):
                ov.cell(row=r, column=c).border = box
                if c > 2:
                    ov.cell(row=r, column=c).alignment = center
                    ov.cell(row=r, column=c).font = body_font
            r += 1
        if r > block_start:
            # ---- job card units summary (a unit passing N stages is still 1 unit) ----
            units_total = len({iid for s in STAGES
                               for iid in (stage_frames[s][1]["InsulatorID"]
                                           [stage_frames[s][1]["JobCard"] == jc].tolist()
                                           if not stage_frames[s][1].empty else [])})
            row_units, row_rej, row_net = r, r + 1, r + 2
            ov.cell(row=row_units, column=1, value=jc).font = sub_font
            ov.cell(row=row_units, column=2,
                    value="TOTAL UNITS IN PRODUCTION").font = sub_font
            ov.cell(row=row_units, column=3, value=units_total).font = sub_font
            ov.cell(row=row_rej, column=1, value=jc).font = sub_font
            ov.cell(row=row_rej, column=2, value="REJECTED UNITS").font = sub_font
            ov.cell(row=row_rej, column=3,
                    value=f"=SUM(F{block_start}:F{r-1})").font = sub_font
            ov.cell(row=row_net, column=1, value=jc).font = sub_font
            ov.cell(row=row_net, column=2,
                    value="NET PRODUCED (Units − Rejected)").font = sub_font
            ov.cell(row=row_net, column=3,
                    value=f"=C{row_units}-C{row_rej}").font = sub_font
            for rw in (row_units, row_rej, row_net):
                for c in range(1, 7):
                    ov.cell(row=rw, column=c).border = box
                    ov.cell(row=rw, column=c).fill = sub_fill
                ov.cell(row=rw, column=3).alignment = center
            r = row_net + 2                          # blank spacer between job cards
    if not all_jcs:
        ov.cell(row=r, column=1,
                value="No records in the selected shift window.").font = body_font
        r += 1
    ov.cell(row=r, column=1, value="Reused Rejected (flagged)").font = sub_font
    ov.cell(row=r, column=2, value=len(reused_win)).font = sub_font
    for w_, col in zip([16, 20, 16, 10, 10, 12], "ABCDEF"):
        ov.column_dimensions[col].width = w_
    ov.freeze_panes = "A5"

    # ---- Stage sheets ---------------------------------------------------
    for s in STAGES:
        summ, det = stage_frames[s]
        n_shift_rows = (det[det["Shift Date"] != ""]
                        .groupby(["Shift Date", "Shift", "JobCard", "MRNo"])
                        .ngroups if not det.empty else 0)
        DETAIL_START = max(100, len(summ) + n_shift_rows + 22)
        ws = wb.create_sheet(STAGE_LABELS[s][:31])
        sheet_title(ws, f"{STAGE_LABELS[s]} — Production Entry", win_txt)

        ws["A4"] = "SUMMARY (Job Card + MR level — enter these in ERP)"
        ws["A4"].font = sub_font; ws["A4"].fill = sub_fill
        heads = ["Job Card", "MR No", "Total Produced Qty", "OK Qty",
                 "Rework Qty", "Rejection Qty"]
        for c, h in enumerate(heads, 1):
            ws.cell(row=5, column=c, value=h)
        style_header(ws, 5, len(heads))

        rr = 6
        if not summ.empty:
            D = DETAIL_START
            for _, row in summ.iterrows():
                ws.cell(row=rr, column=1, value=row["Job Card"]).font = body_font
                ws.cell(row=rr, column=2, value=row["MR No"]).font = body_font
                # live COUNTIFS against the detail table below (JobCard=B, MR=C, Status=E)
                ws.cell(row=rr, column=3,
                        value=f'=COUNTIFS($B${D}:$B$100000,A{rr},$C${D}:$C$100000,B{rr})')
                ws.cell(row=rr, column=4,
                        value=f'=COUNTIFS($B${D}:$B$100000,A{rr},$C${D}:$C$100000,B{rr},$E${D}:$E$100000,"OK")')
                ws.cell(row=rr, column=5,
                        value=f'=COUNTIFS($B${D}:$B$100000,A{rr},$C${D}:$C$100000,B{rr},$E${D}:$E$100000,"Rework")')
                ws.cell(row=rr, column=6,
                        value=f'=COUNTIFS($B${D}:$B$100000,A{rr},$C${D}:$C$100000,B{rr},$E${D}:$E$100000,"Rejection")')
                for c in range(1, 7):
                    ws.cell(row=rr, column=c).border = box
                    if c > 2:
                        ws.cell(row=rr, column=c).alignment = center
                        ws.cell(row=rr, column=c).font = body_font
                rr += 1
            ws.cell(row=rr, column=1, value="TOTAL").font = sub_font
            for c in range(3, 7):
                col = get_column_letter(c)
                cell = ws.cell(row=rr, column=c, value=f"=SUM({col}6:{col}{rr-1})")
                cell.font = sub_font; cell.alignment = center; cell.border = box
            ws.cell(row=rr, column=1).border = box
            ws.cell(row=rr, column=2).border = box
        else:
            ws.cell(row=rr, column=1, value="No records in the selected shift window.").font = body_font

        # ---- SHIFT-WISE SUMMARY (Date x Shift, merged date cells) ----
        if not det.empty:
            b_start = time((shift_a_start.hour + 12) % 24, shift_a_start.minute)
            a_txt = (f"Shift A  {shift_a_start:%H:%M}–{b_start:%H:%M}")
            b_txt = (f"Shift B  {b_start:%H:%M}–{shift_a_start:%H:%M} (+1 day)")
            srow = rr + 3
            hdr = ws.cell(row=srow, column=1,
                          value="SHIFT-WISE SUMMARY (Date × Shift)")
            hdr.font = sub_font; hdr.fill = sub_fill
            srow += 1
            sheads = ["Date", "Shift", "Job Card", "MR No",
                      "Total Produced Qty", "OK Qty", "Rework Qty", "Rejection Qty"]
            for c, h in enumerate(sheads, 1):
                ws.cell(row=srow, column=c, value=h)
            style_header(ws, srow, len(sheads))
            srow += 1
            D = DETAIL_START
            grouped = (det[det["Shift Date"] != ""]
                       .groupby(["Shift Date", "Shift", "JobCard", "MRNo"])
                       .size().reset_index(name="_n"))
            grouped["_d"] = pd.to_datetime(grouped["Shift Date"], format="%d/%m/%Y")
            grouped = grouped.sort_values(["_d", "Shift", "JobCard", "MRNo"])
            for dt_txt, ddf in grouped.groupby("Shift Date", sort=False):
                d0 = srow
                for _, g in ddf.iterrows():
                    sh = g["Shift"]
                    ws.cell(row=srow, column=2,
                            value=a_txt if sh == "A" else b_txt).font = body_font
                    ws.cell(row=srow, column=3, value=g["JobCard"]).font = body_font
                    ws.cell(row=srow, column=4, value=g["MRNo"]).font = body_font
                    base = (f'$B${D}:$B$100000,C{srow},$C${D}:$C$100000,D{srow},'
                            f'$N${D}:$N$100000,A{d0},$O${D}:$O$100000,"{sh}"')
                    ws.cell(row=srow, column=5, value=f'=COUNTIFS({base})')
                    ws.cell(row=srow, column=6,
                            value=f'=COUNTIFS({base},$E${D}:$E$100000,"OK")')
                    ws.cell(row=srow, column=7,
                            value=f'=COUNTIFS({base},$E${D}:$E$100000,"Rework")')
                    ws.cell(row=srow, column=8,
                            value=f'=COUNTIFS({base},$E${D}:$E$100000,"Rejection")')
                    for c in range(1, 9):
                        ws.cell(row=srow, column=c).border = box
                        if c >= 5:
                            ws.cell(row=srow, column=c).alignment = center
                            ws.cell(row=srow, column=c).font = body_font
                    srow += 1
                # merged date cell spanning this date's shift rows
                ws.cell(row=d0, column=1, value=dt_txt)
                if srow - 1 > d0:
                    ws.merge_cells(start_row=d0, start_column=1,
                                   end_row=srow - 1, end_column=1)
                mc = ws.cell(row=d0, column=1)
                mc.font = sub_font; mc.alignment = Alignment(
                    horizontal="center", vertical="center"); mc.fill = sub_fill

        ws.cell(row=DETAIL_START - 2, column=1,
                value="DETAIL (Insulator level)").font = sub_font
        ws.cell(row=DETAIL_START - 2, column=1).fill = sub_fill
        dheads = ["Insulator ID", "Job Card", "MR No", "SKU", "Status",
                  "Prod Machine", "Prod Date", "Prod Time",
                  "Rework Machine", "Rework Date/Time",
                  "Rejection Machine", "Rejection Date/Time",
                  "Rejection Reason / Remarks", "Shift Date", "Shift"]
        for c, h in enumerate(dheads, 1):
            ws.cell(row=DETAIL_START - 1, column=c, value=h)
        style_header(ws, DETAIL_START - 1, len(dheads))

        def fmt_dt(ts, f="%d/%m/%Y %H:%M"):
            return ts.strftime(f) if pd.notna(ts) else ""

        rr = DETAIL_START
        if not det.empty:
            for jc, grp_df in det.groupby("JobCard", sort=True):
                band = ws.cell(row=rr, column=1, value=f"Job Card: {jc}")
                band.font = sub_font
                for c in range(1, len(dheads) + 1):
                    ws.cell(row=rr, column=c).fill = sub_fill
                    ws.cell(row=rr, column=c).border = box
                rr += 1
                for _, row in grp_df.iterrows():
                    reason = row["Rejection Reason"]
                    remarks = row["Remarks"]
                    reason_txt = " — ".join(x for x in [reason, remarks] if x)
                    vals = [row["InsulatorID"], row["JobCard"], row["MRNo"],
                            row["SKU"], row["Status"], row["Prod Machine"] or "—",
                            fmt_dt(row["Prod DateTime"], "%d/%m/%Y"),
                            fmt_dt(row["Prod DateTime"], "%H:%M:%S"),
                            row["Rework Machine"] or "",
                            fmt_dt(row["Rework DateTime"]),
                            row["Rejection Machine"] or "",
                            fmt_dt(row["Rejection DateTime"]),
                            reason_txt, row["Shift Date"], row["Shift"]]
                    for c, v in enumerate(vals, 1):
                        cell = ws.cell(row=rr, column=c, value=v)
                        cell.font = body_font; cell.border = box
                        if c in (5, 6, 7, 8, 9, 10, 11, 12):
                            cell.alignment = center
                    st_ = row["Status"]
                    if st_ == "Rejection":
                        ws.cell(row=rr, column=5).font = Font(name="Arial", size=10,
                                                              bold=True, color="C0392B")
                    elif st_ == "Rework":
                        ws.cell(row=rr, column=5).font = Font(name="Arial", size=10,
                                                              bold=True, color="B9770E")
                    rr += 1
                rr += 1

        # ---- REWORK & REJECTION DETAILS section ----
        rrej = det[det["Status"].isin(["Rework", "Rejection"])] if not det.empty \
            else pd.DataFrame()
        if not rrej.empty:
            rr += 1
            hdr = ws.cell(row=rr, column=1,
                          value="REWORK & REJECTION DETAILS — reasons and machines")
            hdr.font = sub_font; hdr.fill = sub_fill
            rr += 1
            rheads = ["Insulator ID", "Type", "Machine", "Stage", "Job Card",
                      "MR No", "Date/Time", "Reason / Remarks"]
            for c, h in enumerate(rheads, 1):
                ws.cell(row=rr, column=c, value=h)
            style_header(ws, rr, len(rheads))
            rr += 1
            for _, row in rrej.sort_values(["Status", "InsulatorID"]).iterrows():
                is_rej = row["Status"] == "Rejection"
                machine = (row["Rejection Machine"] if is_rej
                           else row["Rework Machine"]) or "—"
                ts = (row["Rejection DateTime"] if is_rej
                      else row["Rework DateTime"])
                reason = " — ".join(x for x in
                                    [row["Rejection Reason"] if is_rej else "",
                                     row["Remarks"]] if x) or "—"
                vals = [row["InsulatorID"], row["Status"], machine,
                        STAGE_LABELS[s], row["JobCard"], row["MRNo"],
                        fmt_dt(ts), reason]
                for c, v in enumerate(vals, 1):
                    cell = ws.cell(row=rr, column=c, value=v)
                    cell.font = body_font; cell.border = box
                    if c in (2, 3, 4, 7):
                        cell.alignment = center
                ws.cell(row=rr, column=2).font = Font(
                    name="Arial", size=10, bold=True,
                    color="C0392B" if is_rej else "B9770E")
                rr += 1
        for w_, col in zip([22, 13, 15, 11, 11, 12, 11, 9, 13, 16, 15, 17, 34, 12, 7],
                           ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O"]):
            ws.column_dimensions[col].width = w_
        ws.freeze_panes = "A6"

    # ---- Reused Rejected sheet ------------------------------------------
    ws = wb.create_sheet("Reused Rejected")
    sheet_title(ws, "Reused Rejected Quantity — rejected units re-run from Priming (same QR)",
                win_txt)
    ws["A3"] = ("These units carry an OLD Rejection timestamp but NEWER production "
                "timestamps — the rubber was peeled and the rod restarted from Priming.")
    ws["A3"].font = Font(name="Arial", size=9, italic=True, color="6B7280")
    heads = ["Insulator ID", "Job Card", "MR No", "SKU", "Rejected At Stage",
             "Rejection Machine", "Rejection Reason", "Rejection Date/Time",
             "Reused From Stage", "Reuse Date/Time", "Latest Stage Reached"]
    for c, h in enumerate(heads, 1):
        ws.cell(row=5, column=c, value=h)
    style_header(ws, 5, len(heads))
    rr = 6
    if not reused_win.empty:
        for _, row in reused_win.iterrows():
            vals = [row["InsulatorID"], row["JobCard"], row["MRNo"], row["SKU"],
                    row["Rejected At Stage"], row["Rejection Machine"],
                    row["Rejection Reason"],
                    row["Rejection DateTime"].strftime("%d/%m/%Y %H:%M"),
                    row["Reused From Stage"],
                    row["Reuse DateTime"].strftime("%d/%m/%Y %H:%M"),
                    row["Latest Stage Reached"]]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=rr, column=c, value=v)
                cell.font = body_font; cell.border = box
            rr += 1
        ws.cell(row=rr + 1, column=1, value="Reused Rejected Quantity:").font = sub_font
        ws.cell(row=rr + 1, column=2, value=f"=COUNTA(A6:A{rr-1})").font = sub_font
    else:
        ws.cell(row=6, column=1,
                value="No reused-rejected units detected in the selected window.").font = body_font
    for w_, col in zip([22, 14, 16, 12, 16, 16, 22, 18, 16, 18, 18],
                       ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]):
        ws.column_dimensions[col].width = w_

    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


# ======================================================================
# Charts (Altair — bundled with Streamlit, no extra install needed)
# ======================================================================
def render_charts(stage_frames, reused_all):
    import altair as alt

    status_scale = alt.Scale(domain=["OK", "Rework", "Rejection"],
                             range=["#2E8B57", "#E8A317", "#C0392B"])
    all_det = pd.concat(
        [d.assign(Stage=STAGE_LABELS[s]) for s, (m, d) in stage_frames.items()
         if not d.empty], ignore_index=True) \
        if any(not d.empty for _, d in stage_frames.values()) else pd.DataFrame()
    if all_det.empty:
        st.info("No scan records in the selected window — nothing to chart.")
        return

    st.markdown("#### Filters")
    f1, f2, f3 = st.columns(3)
    jcs = sorted(all_det["JobCard"].dropna().unique())
    stages_avail = [STAGE_LABELS[s] for s in STAGES if not stage_frames[s][1].empty]
    sel_jc = f1.multiselect("Job Card", jcs, default=jcs)
    sel_st = f2.multiselect("Stage", stages_avail, default=stages_avail)
    sel_status = f3.multiselect("Status", ["OK", "Rework", "Rejection"],
                                default=["OK", "Rework", "Rejection"])
    fdf = all_det[all_det["JobCard"].isin(sel_jc) & all_det["Stage"].isin(sel_st)
                  & all_det["Status"].isin(sel_status)].copy()
    if fdf.empty:
        st.warning("No data after filters."); return
    st.caption("Use the **⋯ menu on any chart → 'Save as PNG / SVG'** to export it "
               "as an image.")

    stage_order = [STAGE_LABELS[s] for s in STAGES]

    def labelled_bar(data, x, color=None, x_order=None, title="", stacked=True):
        base = alt.Chart(data).encode(
            x=alt.X(f"{x}:N", sort=x_order, title=None,
                    axis=alt.Axis(labelAngle=0)),
            y=alt.Y("Qty:Q", title="Quantity"),
            tooltip=[x, "Status", "Qty"] if color else [x, "Qty"])
        if color:
            bars = base.mark_bar().encode(
                color=alt.Color("Status:N", scale=status_scale, title=None),
                **({} if stacked else {"xOffset": "Status:N"}))
            text = base.mark_text(dy=-6, fontWeight="bold").encode(
                text="Qty:Q", **({} if stacked else {"xOffset": "Status:N"}),
                y=alt.Y("Qty:Q", stack="zero" if stacked else None))
        else:
            bars = base.mark_bar(color=DECCAN_BLUE)
            text = base.mark_text(dy=-6, fontWeight="bold").encode(text="Qty:Q")
        return (bars + text).properties(title=title, height=320)

    c1, c2 = st.columns(2)
    with c1:
        g = fdf.groupby(["Stage", "Status"]).size().reset_index(name="Qty")
        st.altair_chart(labelled_bar(g, "Stage", color=True, x_order=stage_order,
                                     title="Stage-wise Production (OK / Rework / Rejection)"),
                        width="stretch")
    with c2:
        g = fdf.groupby(["JobCard", "Status"]).size().reset_index(name="Qty")
        st.altair_chart(labelled_bar(g, "JobCard", color=True, stacked=False,
                                     title="Job Card–wise Quantities"),
                        width="stretch")

    c3, c4 = st.columns(2)
    with c3:
        mdf = fdf.assign(Machine=fdf["Prod Machine"].where(
            fdf["Status"] == "OK",
            fdf["Rework Machine"].where(fdf["Status"] == "Rework",
                                        fdf["Rejection Machine"])))
        mdf = mdf[mdf["Machine"].astype(str).str.strip().ne("") &
                  mdf["Machine"].ne("—")]
        if not mdf.empty:
            g = mdf.groupby(["Machine", "Status"]).size().reset_index(name="Qty")
            st.altair_chart(labelled_bar(g, "Machine", color=True,
                                         title="Machine-wise Output"),
                            width="stretch")
    with c4:
        okd = fdf[pd.notna(fdf["Prod DateTime"])].copy()
        if not okd.empty:
            okd["Hour"] = okd["Prod DateTime"].dt.floor("h")
            g = okd.groupby(["Hour", "Stage"]).size().reset_index(name="Qty")
            line = alt.Chart(g).mark_line(point=True).encode(
                x=alt.X("Hour:T", title=None),
                y=alt.Y("Qty:Q", title="Units scanned"),
                color=alt.Color("Stage:N", title=None),
                tooltip=["Hour:T", "Stage", "Qty"]).properties(
                title="Hourly Throughput by Stage", height=320)
            st.altair_chart(line, width="stretch")

    c5, c6 = st.columns(2)
    with c5:
        rank = {"OK": 0, "Rework": 1, "Rejection": 2}
        ustat = fdf.groupby("InsulatorID")["Status"].agg(
            lambda g: max(g, key=lambda s: rank[s]))
        g = ustat.value_counts().rename_axis("Status").reset_index(name="Qty")
        donut = alt.Chart(g).mark_arc(innerRadius=70).encode(
            theta="Qty:Q", color=alt.Color("Status:N", scale=status_scale),
            tooltip=["Status", "Qty"]).properties(
            title=f"Unit Quality Split — {ustat.size} distinct insulators "
                  "(OK = clean through all stages)", height=320)
        st.altair_chart(donut, width="stretch")
    with c6:
        if not reused_all.empty:
            g = reused_all.groupby("Rejected At Stage").size().reset_index(name="Qty")
            st.altair_chart(labelled_bar(g, "Rejected At Stage",
                                         title="Reused Rejected Units by Rejection Stage"),
                            width="stretch")
        else:
            st.info("No reused-rejected units in the uploaded data.")


# ======================================================================
# PRODUCTION ENTRY VISUALIZER — page
# ======================================================================
def page_visualizer():
    page_header("Daily Production Entry Visualizer",
                "Upload the ScanLog export, pick the shift window, and generate "
                "the ERP-ready stage-wise production entry report.")

    up = st.file_uploader("Upload ScanLog export", type=["xlsx", "xls", "csv", "iqy"],
                          help="Excel or CSV export of the ScanLog SharePoint list. "
                               ".iqy files are detected and explained.")
    if up is None:
        st.info("Waiting for a file. In SharePoint: **Export → Export to CSV** "
                "(or open in Excel and save as .xlsx), then upload here.")
        return

    with st.expander("Data-repair options", expanded=False):
        do_shift = st.checkbox("Repair column-shifted rows (Rejection_MC holding a timestamp)",
                               value=True)
        do_md = st.checkbox("Auto-correct swapped month/day timestamps "
                            "(e.g. 2026-10-06 that should be 2026-06-10)", value=True)
        do_dedup = st.checkbox("Detect duplicate/test rows (identical timestamps to "
                               "another row)", value=True)

    raw, err = load_scan_export(up)
    if err:
        (st.warning if up.name.lower().endswith(".iqy") else st.error)(err)
        return

    df_all, notes, dup_df = prepare_data(raw, do_shift_repair=do_shift, do_md_fix=do_md,
                                         do_dedup=do_dedup)
    for n in notes:
        st.info("🛠 " + n)

    # ---- exclusion control: prefilled with detected duplicates, fully editable ----
    every_id = sorted(pd.concat([df_all["Title"],
                                 dup_df["Title"] if not dup_df.empty else pd.Series(dtype=str)])
                      .dropna().astype(str).unique())
    detected = sorted(dup_df["Title"].tolist()) if not dup_df.empty else []
    excl = st.multiselect(
        "Excluded insulators (quarantined rows pre-selected — remove an ID from "
        "this box to re-include it as a real unit)",
        every_id, default=detected)
    if not dup_df.empty:
        with st.expander(f"Why these {len(dup_df)} rows were quarantined"):
            st.dataframe(dup_df[["Title", "FlagReason", "JobCard", "MRNo"]],
                         width="stretch", hide_index=True)
            st.caption("These rows exist in the ScanLog SharePoint list itself — "
                       "delete them there to clean this up permanently.")
    full = pd.concat([df_all, dup_df.drop(columns=["FlagReason"], errors="ignore")],
                     ignore_index=True) if not dup_df.empty else df_all
    full.attrs["reason_col"] = df_all.attrs.get("reason_col")
    df = full[~full["Title"].isin(excl)].copy()
    df.attrs["reason_col"] = df_all.attrs.get("reason_col")

    # ---- machine -> process mapping (editable; used when StageAsked is blank) ----
    mmap = infer_machine_map(df)
    with st.expander("Machine → Process mapping (used to attribute Rework/Rejection "
                     "when StageAsked is blank)", expanded=False):
        st.caption("Prefilled from this file's stage machine columns + Deccan machine "
                   "naming (FP-145*=CBM, M*=Moulding, FP-165*=CAM, SML*=SML, "
                   "C*=C-Frame). Edit freely, or upload a mapping file with columns "
                   "**Machine, Process**.")
        map_up = st.file_uploader("Optional mapping file", type=["xlsx", "csv"],
                                  key="mapfile")
        if map_up is not None:
            try:
                mf = (pd.read_csv(map_up) if map_up.name.lower().endswith(".csv")
                      else pd.read_excel(map_up))
                cols = {c.lower().strip(): c for c in mf.columns}
                mc, pc = cols.get("machine"), cols.get("process") or cols.get("stage")
                if mc and pc:
                    for _, r in mf.iterrows():
                        if pd.notna(r[mc]) and str(r[pc]).strip() in STAGES:
                            mmap[str(r[mc]).strip()] = str(r[pc]).strip()
                    st.success(f"Loaded {len(mf)} mapping rows.")
                else:
                    st.error("Mapping file needs 'Machine' and 'Process' columns.")
            except Exception as e:
                st.error(f"Could not read mapping file: {e}")
        map_df = pd.DataFrame(sorted(mmap.items()), columns=["Machine", "Process"])
        edited = st.data_editor(map_df, num_rows="dynamic", width="stretch",
                                column_config={"Process": st.column_config.SelectboxColumn(
                                    options=STAGES)},
                                key="mapeditor")
        mmap = {str(r["Machine"]).strip(): r["Process"]
                for _, r in edited.iterrows()
                if pd.notna(r["Machine"]) and r["Process"] in STAGES}

    all_dt_cols = [c for c in df.columns if c.endswith("_DT")]
    all_ts = pd.concat([df[c] for c in all_dt_cols]).dropna()
    if all_ts.empty:
        st.error("No timestamps found in the file."); return
    dmin, dmax = all_ts.min().date(), all_ts.max().date()

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Valid insulators", f"{df['Title'].nunique()}",
              delta=f"-{len(excl)} excluded" if excl else None, delta_color="off")
    m2.metric("Job Cards", f"{df['JobCard'].nunique() if 'JobCard' in df else '—'}")
    m3.metric("First scan", dmin.strftime("%d/%m/%Y"))
    m4.metric("Last scan", dmax.strftime("%d/%m/%Y"))

    st.markdown("### Shift window")
    st.caption("Default covers the FULL file so every stage shows its data. For one "
               "day's ERP entry select e.g. 6th → 7th = everything scanned between "
               "6th 08:00 and 7th 08:00. A stage sheet only lists units whose scan "
               "at THAT stage falls inside the window.")
    c1, c2 = st.columns([3, 1])
    rng = c1.date_input("Production date range",
                        value=(dmin, dmax + timedelta(days=1)),
                        min_value=dmin - timedelta(days=1),
                        max_value=dmax + timedelta(days=1), format="DD/MM/YYYY")
    shift_start = c2.time_input("Shift start time", value=time(8, 0),
                                help="Window runs from this time on the first date "
                                     "to this time on the last date.")
    if isinstance(rng, tuple) and len(rng) == 2:
        d_from, d_to = rng
    else:
        d_from = d_to = rng if isinstance(rng, date) else rng[0]
    win_start, win_end = shift_window(d_from, d_to, shift_start)
    st.markdown(f"**Window:** {win_start:%d/%m/%Y %H:%M} → {win_end:%d/%m/%Y %H:%M}")

    stage_frames, reused_win, reused_all, unattr = build_stage_frames(
        df, win_start, win_end, machine_map=mmap, shift_a_start=shift_start)

    if not unattr.empty:
        st.warning(f"⚠ {len(unattr)} Rework/Rejection scan(s) could not be attributed "
                   "to a stage (StageAsked blank and machine not in the mapping). "
                   "Add those machines to the mapping above.")
        st.dataframe(unattr, width="stretch", hide_index=True)

    unit_status = unit_level_status(stage_frames)
    distinct_units = len(unit_status)
    rjq = sum(1 for v in unit_status.values() if v == "Rejection")
    rwq = sum(1 for v in unit_status.values() if v == "Rework")
    okq = distinct_units - rjq - rwq
    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("Distinct units", distinct_units)
    k2.metric("OK units (no rework/rejection)", okq)
    k3.metric("Reworked units", rwq)
    k4.metric("Rejected units", rjq)
    k5.metric("Reused Rejected", len(reused_win))

    if distinct_units == 0 and reused_win.empty:
        st.warning("No scans fall inside this window — adjust the dates or shift time.")
        return

    st.markdown("### ERP-ready Excel report")
    xlsx = write_excel_report(stage_frames, reused_win, win_start, win_end,
                              shift_a_start=shift_start)
    st.download_button(
        "⬇ Download Production Entry Report (.xlsx)", data=xlsx,
        file_name=f"Deccan_ProductionEntry_{win_start:%d%m%Y}_{win_end:%d%m%Y}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary", width="stretch")

    st.markdown("### Preview")
    tabs = st.tabs([STAGE_LABELS[s] for s in STAGES] + ["Reused Rejected"])
    for tab, s in zip(tabs[:-1], STAGES):
        with tab:
            summ, det = stage_frames[s]
            if det.empty:
                st.caption("No scans at this stage in the selected window.")
                continue
            st.markdown("**Summary — enter these numbers in the ERP**")
            st.dataframe(summ, width="stretch", hide_index=True)
            sdet = det[det["Shift Date"] != ""]
            if not sdet.empty and sdet["Shift Date"].nunique() * \
                    sdet["Shift"].nunique() > 1:
                st.markdown("**Shift-wise summary (Date × Shift)**")
                sg = (sdet.groupby(["Shift Date", "Shift", "JobCard", "MRNo"])
                      ["Status"].agg(Total="size",
                                     OK=lambda g: int((g == "OK").sum()),
                                     Rework=lambda g: int((g == "Rework").sum()),
                                     Rejection=lambda g: int((g == "Rejection").sum()))
                      .reset_index())
                sg["_d"] = pd.to_datetime(sg["Shift Date"], format="%d/%m/%Y")
                sg = sg.sort_values(["_d", "Shift"]).drop(columns="_d")
                sg["Shift"] = sg["Shift"].map({"A": "Shift A", "B": "Shift B"})
                st.dataframe(sg, width="stretch", hide_index=True)
            st.markdown("**Detail — insulator level**")
            show = det.copy()
            for col in ["Prod DateTime", "Rework DateTime", "Rejection DateTime"]:
                show[col] = show[col].apply(
                    lambda t: t.strftime("%d/%m/%Y %H:%M:%S") if pd.notna(t) else "")
            st.dataframe(show[["InsulatorID", "JobCard", "MRNo", "SKU", "Status",
                               "Prod Machine", "Prod DateTime", "Rework Machine",
                               "Rework DateTime", "Rejection Machine",
                               "Rejection DateTime", "Rejection Reason", "Remarks"]],
                         width="stretch", hide_index=True)
            rrej = show[show["Status"].isin(["Rework", "Rejection"])]
            if not rrej.empty:
                st.markdown("**Rework & Rejection details — reasons and machines**")
                rt = pd.DataFrame({
                    "InsulatorID": rrej["InsulatorID"],
                    "Type": rrej["Status"],
                    "Machine": rrej.apply(
                        lambda r: (r["Rejection Machine"] if r["Status"] == "Rejection"
                                   else r["Rework Machine"]) or "—", axis=1),
                    "Stage": STAGE_LABELS[s],
                    "Job Card": rrej["JobCard"], "MR No": rrej["MRNo"],
                    "Date/Time": rrej.apply(
                        lambda r: (r["Rejection DateTime"] if r["Status"] == "Rejection"
                                   else r["Rework DateTime"]), axis=1),
                    "Reason / Remarks": rrej.apply(
                        lambda r: " — ".join(x for x in
                                             [r["Rejection Reason"]
                                              if r["Status"] == "Rejection" else "",
                                              r["Remarks"]] if x) or "—", axis=1),
                })
                st.dataframe(rt, width="stretch", hide_index=True)
    with tabs[-1]:
        if reused_win.empty:
            st.caption("No reused-rejected units detected in this window.")
        else:
            show = reused_win.copy()
            show["Rejection DateTime"] = show["Rejection DateTime"].dt.strftime("%d/%m/%Y %H:%M")
            show["Reuse DateTime"] = show["Reuse DateTime"].dt.strftime("%d/%m/%Y %H:%M")
            st.dataframe(show, width="stretch", hide_index=True)

    st.markdown("### Analytics")
    if st.toggle("📊 Show production charts", value=False,
                 help="Interactive dashboard for the selected window. Charts export "
                      "to PNG/SVG via each chart's ⋯ menu."):
        render_charts(stage_frames, reused_all)


# ======================================================================
# APP SHELL
# ======================================================================
st.set_page_config(page_title="Deccan Dakshata Suite", page_icon="🔳", layout="wide")
inject_css()
sidebar_header()

with st.sidebar:
    page = st.radio("Module", ["🔳 QR Generator", "📊 Production Entry Visualizer"],
                    index=0, label_visibility="collapsed")
    page = page.split(" ", 1)[1]
    st.markdown("---")
    st.markdown(f"<div style='color:{DECCAN_GREY};font-size:12px'>"
                "Deccan Enterprises Pvt. Ltd.<br>Project Dakshata — QR Traceability"
                "</div>", unsafe_allow_html=True)

if page == "QR Generator":
    page_qr_generator()
else:
    page_visualizer()
